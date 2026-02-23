# =============================================================================
# tb_pivot_excel/v1/main.py
#
# Description:
#   Fetches timeseries telemetry from ThingsBoard, builds pivot and aggregated
#   DataFrames, and exports them to a formatted multi-sheet Excel file.
#
#   Functionality includes:
#     - Parses and validates incoming widget payload
#     - Merges user reportConfig with defaults from settings.py
#     - Fetches telemetry from ThingsBoard REST API (with JWT auto-refresh)
#     - Builds raw and pivot DataFrames
#     - Resamples pivot to Daily / Weekly / Monthly / Yearly aggregations
#     - Exports all sheets to .xlsx with formatting, freeze panes, merged headers
#     - Prints effective config to terminal before execution
#
# Sections:
#   1. Config & Merging        _merge_report_config(), _parse_payload()
#   2. ThingsBoard Fetch       _fetch_timeseries(), _to_raw_rows()
#   3. Aggregation Helpers     _get_agg_func(), _week_start_offset(),
#                              _build_agg_dict(), _resample_pivot()
#   4. Excel Formatting        _get_fill_color(), _resolve_headers(),
#                              _write_headers(), _autosize(), _format_sheet()
#   5. Core Logic              build_dataframes_from_widget_payload(),
#                              generate_xlsx_from_widget_payload()
#   6. Public API              generate_pivot_excel_file()
#   7. CLI                     _read_json(), main()
#
# Version: 2026.02.00 - Wit Wonghanchao
#   - Refactored settings to DEFAULT_REPORT_CONFIG mirroring reportConfig
#   - Replaced 4 merge functions with single _merge_report_config()
#   - Added effective config printout at start of execution
#   - Reorganized code into logical sections for maintainability
# =============================================================================

import json
import os
import re
import sys
from datetime import datetime

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import Config
from utils.tb_auth import get_headers, get_jwt
import projects.tb_pivot_excel.v1.settings as S


# ── 1. Config & Merging ───────────────────────────────────────────────────────

def _merge_report_config(payload_report_config: dict) -> dict:
    return S.resolve_config(S.DEFAULT_REPORT_CONFIG, payload_report_config)


def _parse_payload(payload: dict):
    tz       = payload.get("timezone") or S.DEFAULT_TIMEZONE
    te       = payload.get("timeEpoch") or {}
    start_ts = int(te.get("startTs_ms"))
    end_ts   = int(te.get("endTs_ms"))
    entities = payload.get("entities") or []
    keys     = payload.get("keys") or []
    q        = payload.get("query") or {}
    agg      = q.get("agg") or "NONE"
    interval = q.get("interval")                          # ← NEW
    limit    = int(q.get("limit") or S.MAX_POINTS_PER_KEY_PER_ENTITY)
    order    = (q.get("order") or "ASC").upper()

    if len(entities) > S.MAX_ENTITIES:
        entities = entities[:S.MAX_ENTITIES]
    if len(keys) > S.MAX_KEYS:
        keys = keys[:S.MAX_KEYS]
    if limit > S.MAX_POINTS_PER_KEY_PER_ENTITY:
        limit = S.MAX_POINTS_PER_KEY_PER_ENTITY

    norm_entities = []
    for e in entities:
        et   = (e.get("type") or "ASSET").upper()
        eid  = e.get("id")
        name = e.get("name") or eid
        if eid:
            norm_entities.append({"type": et, "id": eid, "name": name})

    norm_keys = [str(k) for k in keys if k is not None and str(k).strip()]

    rc = _merge_report_config(payload.get("reportConfig"))

    result = {
        "timezone":           tz,
        "startTs":            start_ts,
        "endTs":              end_ts,
        "entities":           norm_entities,
        "keys":               norm_keys,
        "agg":                agg,
        "interval":           interval,                   # ← NEW
        "limit":              limit,
        "order":              order,
        "filename":           rc["filename"],
        "filename_timestamp": rc["filename_timestamp"],
        "fmt":                rc["formatting"],
        "column_map":         rc["column_map"],
        "agg_map":            rc["agg_map"],
        "sheets_cfg":         rc["sheets"],
    }

    return result


# ── 2. ThingsBoard Fetch ──────────────────────────────────────────────────────

_MAX_INTERVALS_PER_REQUEST = 700   # conservative ThingsBoard server cap


def _fetch_timeseries_single(tb_url, tenant_id, entity_type, entity_id,
                             keys, start_ts, end_ts, limit, agg, interval_param):
    """One raw HTTP request; handles 401 token refresh."""
    url = (
        f"{tb_url}/api/plugins/telemetry/{entity_type}/{entity_id}/values/timeseries"
        f"?keys={','.join(keys)}&startTs={start_ts}&endTs={end_ts}&limit={limit}"
        f"&agg={agg}{interval_param}&useStrictDataTypes=true"
    )
    r = requests.get(url, headers=get_headers(tenant_id), timeout=60)
    if r.status_code == 401:
        base_dir   = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
        cache_path = os.path.join(base_dir, ".cache", f"tb_token_{tenant_id}.json")
        try:
            os.remove(cache_path)
        except Exception:
            pass
        _ = get_jwt(tenant_id)
        r = requests.get(url, headers=get_headers(tenant_id), timeout=60)
    if not r.ok:
        print(f"[TB ERROR {r.status_code}] {r.text[:500]}", file=sys.stderr)
    r.raise_for_status()
    return r.json()


def _fetch_timeseries(tb_url, tenant_id, entity_type, entity_id,
                      keys, start_ts, end_ts, limit, agg, interval=None):
    interval_param = f"&interval={interval}&intervalType=MILLISECONDS" if interval and agg != "NONE" else ""

    # Chunk when aggregated interval would exceed server cap
    if interval and agg != "NONE":
        n_intervals = (end_ts - start_ts) / interval
        if n_intervals > _MAX_INTERVALS_PER_REQUEST:
            chunk_ms = int(_MAX_INTERVALS_PER_REQUEST * interval)
            merged: dict = {}
            t = start_ts
            while t < end_ts:
                chunk_end = min(t + chunk_ms, end_ts)
                chunk = _fetch_timeseries_single(
                    tb_url, tenant_id, entity_type, entity_id,
                    keys, t, chunk_end, limit, agg, interval_param,
                )
                for key, points in chunk.items():
                    merged.setdefault(key, []).extend(points)
                t = chunk_end
            return merged

    return _fetch_timeseries_single(
        tb_url, tenant_id, entity_type, entity_id,
        keys, start_ts, end_ts, limit, agg, interval_param,
    )


def _to_raw_rows(entity_name, data_dict):
    rows = {}
    for k, points in (data_dict or {}).items():
        for p in points:
            ts = p.get("ts")
            if ts is None:
                continue
            rows.setdefault(ts, {"ts": ts, "entity": entity_name})
            rows[ts][k] = p.get("value")
    return list(rows.values())


# ── 3. Aggregation Helpers ────────────────────────────────────────────────────

def _get_agg_func(key: str, agg_map: dict) -> str:
    return agg_map.get(key) or agg_map.get("default") or "mean"


def _week_start_offset(week_start: str) -> int:
    return 6 if week_start.lower() == "sunday" else 0


def _build_agg_dict(data_cols: list, agg_map: dict) -> dict:
    return {col: _get_agg_func(col, agg_map) for col in data_cols}


def _resample_pivot(df_pivot: pd.DataFrame, freq: str, agg_map: dict,
                    sheets_cfg: dict) -> pd.DataFrame:
    df = df_pivot.copy()
    df = df.set_index("Timestamp")
    data_cols = [c for c in df.columns]

    agg_dict = _build_agg_dict(data_cols, agg_map)

    if freq == "W":
        offset   = _week_start_offset(sheets_cfg.get("week_start", "Sunday"))
        freq_str = "W-SAT" if offset == 6 else "W-SUN"
    else:
        freq_str = freq

    df_resampled = df.resample(freq_str).agg(agg_dict)

    data_start = df_pivot["Timestamp"].min()
    data_end   = df_pivot["Timestamp"].max()

    complete_rows = []
    for period_end, row in df_resampled.iterrows():
        if freq == "D":
            period_start    = period_end.normalize()
            period_end_excl = period_start + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        elif freq == "W":
            period_end_excl = period_end
            period_start    = period_end - pd.Timedelta(days=6)
        elif freq == "MS":
            period_start    = period_end
            next_month      = period_start + pd.offsets.MonthEnd(1)
            period_end_excl = next_month.normalize() + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        elif freq == "YS":
            period_start    = period_end
            period_end_excl = period_start + pd.offsets.YearEnd(1)
        else:
            period_start    = period_end
            period_end_excl = period_end

        first_day_complete = (data_start == data_start.normalize())
        if period_start >= data_start.normalize() and period_end_excl <= data_end:
            if period_start.date() == data_start.date() and not first_day_complete:
                continue
            complete_rows.append((period_start, row))

    if not complete_rows:
        return pd.DataFrame()

    dates, rows = zip(*complete_rows)
    df_out = pd.DataFrame(list(rows), columns=data_cols)
    df_out.insert(0, "Date", [d.date() for d in dates])
    return df_out


# ── 4. Excel Formatting Helpers ───────────────────────────────────────────────

def _get_fill_color(fmt: dict, row_index: int) -> str:
    colors = fmt.get("header_fill_colors", ["B8CCE4", "D9E1F2", "EEF2FA"])
    idx = min(row_index, len(colors) - 1)
    return colors[idx]


def _resolve_headers(pivot_cols: list, column_map: dict) -> tuple:
    resolved = {}
    for col in pivot_cols:
        if col == "Date":
            resolved[col] = ["Date"]
        elif col == "Timestamp":
            resolved[col] = ["Timestamp"]
        elif col in column_map:
            resolved[col] = list(column_map[col])
        else:
            parts = col.split(" ", 1)
            resolved[col] = parts if len(parts) == 2 else [col]

    max_rows = max(len(v) for v in resolved.values())
    for col in resolved:
        while len(resolved[col]) < max_rows:
            resolved[col].append("")

    return resolved, max_rows


def _write_headers(ws, cols: list, fmt: dict, column_map: dict, ts_col: str = "Timestamp"):
    resolved, max_header_rows = _resolve_headers(cols, column_map)

    thin             = Side(style=fmt["border_style"])
    border           = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font      = Font(bold=fmt["header_font_bold"], size=fmt["header_font_size"])
    header_alignment = Alignment(horizontal=fmt["header_alignment"], vertical="center")

    for col_idx, col in enumerate(cols, start=1):
        labels = resolved[col]
        for row_idx, label in enumerate(labels):
            cell           = ws.cell(row=row_idx + 1, column=col_idx)
            cell.value     = label
            fill_color     = _get_fill_color(fmt, row_idx)
            cell.fill      = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font      = header_font
            cell.alignment = header_alignment
            cell.border    = border

    if max_header_rows > 1:
        ts_col_letter = get_column_letter(1)
        ws.merge_cells(f"{ts_col_letter}1:{ts_col_letter}{max_header_rows}")

        for row_idx in range(1):
            excel_row = row_idx + 1
            start_col = 2
            n_cols    = len(cols)
            for col_idx in range(3, n_cols + 2):
                curr_val = ws.cell(excel_row, col_idx).value if col_idx <= n_cols else None
                prev_val = ws.cell(excel_row, start_col).value
                if curr_val != prev_val or col_idx > n_cols:
                    end_col = col_idx - 1
                    if end_col > start_col:
                        ws.merge_cells(
                            start_row=excel_row, start_column=start_col,
                            end_row=excel_row,   end_column=end_col
                        )
                    start_col = col_idx

    return max_header_rows


def _autosize(ws, min_width=18, max_width=60):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _format_sheet(ws, cols: list, fmt: dict, column_map: dict,
                  freeze_key: str = "freeze_pivot", ts_col: str = "Timestamp"):
    thin   = Side(style=fmt["border_style"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # to_excel writes 1 header row; insert extra rows so data is not overwritten
    # when _write_headers produces a multi-row header.
    _, max_header_rows = _resolve_headers(cols, column_map)
    if max_header_rows > 1:
        ws.insert_rows(2, max_header_rows - 1)

    n_header_rows = _write_headers(ws, cols, fmt, column_map, ts_col)

    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(n_header_rows + 1, max_row + 1):
        for col in range(1, max_col + 1):
            cell        = ws.cell(row, col)
            cell.border = border
            if col > 1 and cell.value is not None:
                try:
                    float(cell.value)
                    cell.number_format = fmt["number_format"]
                except (ValueError, TypeError):
                    pass

    freeze = fmt.get(freeze_key, [1, 1])
    ws.freeze_panes = ws.cell(
        row=n_header_rows + freeze[0],
        column=freeze[1] + 1
    ).coordinate

    _autosize(ws, min_width=fmt["min_col_width"], max_width=fmt["max_col_width"])


# ── 5. Core Logic ─────────────────────────────────────────────────────────────

def build_dataframes_from_widget_payload(payload: dict, tenant_id: str, tb_url: str):
    p = _parse_payload(payload)

    raw_rows = []
    for e in p["entities"]:
        data = _fetch_timeseries(
            tb_url=tb_url, tenant_id=tenant_id,
            entity_type=e["type"], entity_id=e["id"],
            keys=p["keys"], start_ts=p["startTs"], end_ts=p["endTs"],
            limit=p["limit"], agg=p["agg"], interval=p.get("interval"),  # ← NEW
        )
        raw_rows.extend(_to_raw_rows(e["name"], data))

    df_raw = pd.DataFrame(raw_rows)
    if df_raw.empty:
        return df_raw, df_raw, {}, p

    # ThingsBoard returns the MIDPOINT of each aggregated interval.
    # Snap ts to interval start for the Pivot sheet, but keep original ts
    # for the Raw Data sheet so it matches the ThingsBoard API response.
    if p["agg"] != "NONE" and p.get("interval"):
        interval_ms = int(p["interval"])
        pivot_ts_ms = df_raw["ts"] - (df_raw["ts"] % interval_ms)
    else:
        pivot_ts_ms = df_raw["ts"]

    # Raw Data: original TB timestamps
    df_raw["Timestamp"] = (
        pd.to_datetime(df_raw["ts"], unit="ms", utc=True)
        .dt.tz_convert(p["timezone"])
        .dt.tz_localize(None)
    )
    df_raw.drop(columns=["ts"], inplace=True)
    df_raw.rename(columns={"entity": "Asset Name"}, inplace=True)

    cols   = ["Timestamp", "Asset Name"] + [k for k in p["keys"] if k in df_raw.columns]
    df_raw = df_raw[cols]

    # Pivot: snapped timestamps (interval start)
    df_pivot_src = df_raw.copy()
    df_pivot_src["Timestamp"] = (
        pd.to_datetime(pivot_ts_ms.values, unit="ms", utc=True)
        .tz_convert(p["timezone"])
        .tz_localize(None)
    )

    df_pivot         = df_pivot_src.pivot_table(index="Timestamp", columns="Asset Name", values=p["keys"], aggfunc="first")
    df_pivot.columns = [f"{asset} {key}" for (key, asset) in df_pivot.columns]
    df_pivot         = df_pivot.sort_index().reset_index()

    col_order = ["Timestamp"] + [c for c in p["column_map"].keys() if c in df_pivot.columns]
    remaining = [c for c in df_pivot.columns if c not in col_order]
    remaining = sorted(remaining, key=lambda c: (c.split(" ", 1)[0], c.split(" ", 1)[1] if " " in c else ""))
    df_pivot  = df_pivot[col_order + remaining]

    if p["order"] == "DESC":
        df_raw   = df_raw.sort_values("Timestamp", ascending=False)
        df_pivot = df_pivot.sort_values("Timestamp", ascending=False)

    agg_map    = p["agg_map"]
    sheets_cfg = p["sheets_cfg"]

    agg_sheets = {}
    for freq, key in [("D", "daily"), ("W", "weekly"), ("MS", "monthly"), ("YS", "yearly")]:
        df_agg = _resample_pivot(df_pivot, freq, agg_map, sheets_cfg)
        if not df_agg.empty:
            agg_sheets[key] = df_agg

    return df_raw, df_pivot, agg_sheets, p


def generate_xlsx_from_widget_payload(payload: dict, tenant_id: str) -> str:
    cfg    = Config()
    from utils.tb_auth import get_tb_url
    tb_url = get_tb_url(tenant_id)

    df_raw, df_pivot, agg_sheets, p = build_dataframes_from_widget_payload(payload, tenant_id, tb_url)

    base_dir     = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
    tenant_cfg   = cfg.get_tenant_config(tenant_id) or {}
    display_name = tenant_cfg.get("display_name", tenant_id)
    folder_name  = re.sub(r'[^\w]+', '_', display_name).strip('_').lower()
    out_dir      = os.path.join(base_dir, "outputs", folder_name)
    os.makedirs(out_dir, exist_ok=True)

    fname = p["filename"]
    if not fname.lower().endswith(".xlsx"):
        fname += ".xlsx"
    if p["filename_timestamp"]:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(out_dir, fname.replace(".xlsx", f"_{ts}.xlsx"))
    else:
        out_path = os.path.join(out_dir, fname)

    fmt        = p["fmt"]
    column_map = p["column_map"]
    pivot_cols = list(df_pivot.columns)

    freq_sheet_map = {
        "daily":   fmt["sheet_daily"],
        "weekly":  fmt["sheet_weekly"],
        "monthly": fmt["sheet_monthly"],
        "yearly":  fmt["sheet_yearly"],
    }

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_pivot.to_excel(writer, index=False, sheet_name=fmt["sheet_pivot"])
        for freq_key, sheet_name in freq_sheet_map.items():
            if freq_key in agg_sheets:
                agg_sheets[freq_key].to_excel(writer, index=False, sheet_name=sheet_name)
        df_raw.to_excel(writer, index=False, sheet_name=fmt["sheet_raw"])

    wb = load_workbook(out_path)

    ws_pivot = wb[fmt["sheet_pivot"]]
    _format_sheet(ws_pivot, pivot_cols, fmt, column_map, "freeze_pivot", "Timestamp")

    agg_cols = ["Date"] + [c for c in pivot_cols if c != "Timestamp"]
    for freq_key, sheet_name in freq_sheet_map.items():
        if freq_key in agg_sheets and sheet_name in wb.sheetnames:
            ws_agg = wb[sheet_name]
            _format_sheet(ws_agg, agg_cols, fmt, column_map, f"freeze_{freq_key}", "Date")

    thin        = Side(style=fmt["border_style"])
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_color  = _get_fill_color(fmt, 0)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(bold=fmt["header_font_bold"], size=fmt["header_font_size"])
    header_aln  = Alignment(horizontal=fmt["header_alignment"], vertical="center")

    ws_raw = wb[fmt["sheet_raw"]]
    for row in range(1, ws_raw.max_row + 1):
        for col in range(1, ws_raw.max_column + 1):
            cell        = ws_raw.cell(row, col)
            cell.border = border
            if row == 1:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = header_aln
            elif col > 1 and cell.value is not None:
                try:
                    float(cell.value)
                    cell.number_format = fmt["number_format"]
                except (ValueError, TypeError):
                    pass

    freeze_raw = fmt["freeze_raw"]
    ws_raw.freeze_panes = ws_raw.cell(
        row=freeze_raw[0] + 1,
        column=freeze_raw[1] + 1
    ).coordinate
    _autosize(ws_raw, min_width=fmt["min_col_width"], max_width=fmt["max_col_width"])

    if S.DEBUG_CONFIG_SHEET:
        ws_cfg = wb.create_sheet("_Config")

        effective_rc = {
            "filename":           p["filename"],
            "filename_timestamp": p["filename_timestamp"],
            "column_map":         p["column_map"],
            "agg_map":            p["agg_map"],
            "sheets":             p["sheets_cfg"],
            "formatting":         p["fmt"],
        }
        debug_rows = [
            ("widget_payload",        json.dumps(payload,       indent=2, ensure_ascii=False)),
            ("effective_reportConfig", json.dumps(effective_rc, indent=2, ensure_ascii=False)),
        ]

        thin        = Side(style=fmt["border_style"])
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_color  = _get_fill_color(fmt, 0)
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        header_font = Font(bold=True, size=fmt["header_font_size"])
        wrap_top    = Alignment(wrap_text=True, vertical="top")

        for c_idx, label in enumerate(["Setting", "Value"], start=1):
            cell           = ws_cfg.cell(row=1, column=c_idx)
            cell.value     = label
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, (key, val) in enumerate(debug_rows, start=2):
            for c_idx, v in enumerate([key, val], start=1):
                cell           = ws_cfg.cell(row=r_idx, column=c_idx)
                cell.value     = v
                cell.border    = border
                cell.alignment = wrap_top

        ws_cfg.column_dimensions["A"].width = 25
        ws_cfg.column_dimensions["B"].width = 80
        ws_cfg.freeze_panes = "A2"
        ws_cfg.row_dimensions[2].height = 200
        ws_cfg.row_dimensions[3].height = 400

    wb.save(out_path)
    return out_path


# ── 6. Public API ─────────────────────────────────────────────────────────────

def generate_pivot_excel_file(payload: dict, tenant_id: str) -> str:
    if payload is None:
        payload = {}
    te = payload.get("timeEpoch") or {}
    if te.get("startTs_ms") is None or te.get("endTs_ms") is None:
        raise ValueError("Missing timeEpoch.startTs_ms or timeEpoch.endTs_ms")
    if not (payload.get("entities") or []):
        raise ValueError("No entities in payload")
    if not (payload.get("keys") or []):
        raise ValueError("No keys in payload")
    return generate_xlsx_from_widget_payload(payload, tenant_id)


# ── 7. CLI ────────────────────────────────────────────────────────────────────

def _read_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    here         = os.path.dirname(__file__)
    payload_path = None
    if len(os.sys.argv) >= 2:
        payload_path = os.sys.argv[1]
    else:
        for cand in ["test_widget_payload.json", "request_example.json"]:
            p = os.path.join(here, cand)
            if os.path.exists(p):
                payload_path = p
                break
    if not payload_path or not os.path.exists(payload_path):
        raise SystemExit("ERROR: No payload JSON found.")
    tenant_id = os.sys.argv[2] if len(os.sys.argv) >= 3 else "lh_production_environment"
    payload   = _read_json(payload_path)
    out       = generate_pivot_excel_file(payload, tenant_id)
    print("output:", out)


if __name__ == "__main__":
    main()
